import configparser
import openpyxl
from stellar_sdk import Server, Keypair, TransactionBuilder, Network, Asset, exceptions
import time

# Read the config file
config = configparser.ConfigParser()
config.read('config.txt')

# Get values from the config file
DISTRIBUTOR_SECRET_KEY = config['DEFAULT']['DISTRIBUTOR_SECRET_KEY']
IS_DEVELOPMENT = config['DEFAULT'].getboolean('IS_DEVELOPMENT')

# Determine the network to use
if IS_DEVELOPMENT:
    network_passphrase = Network.TESTNET_NETWORK_PASSPHRASE
    horizon_url = "https://horizon-testnet.stellar.org"
else:
    network_passphrase = Network.PUBLIC_NETWORK_PASSPHRASE
    horizon_url = "https://horizon.stellar.org"

# Initialize the Stellar SDK
server = Server(horizon_url)

# Load the source account
distributor_keypair = Keypair.from_secret(DISTRIBUTOR_SECRET_KEY)

# Load the Excel file
workbook = openpyxl.load_workbook('database.xlsx')
sheet = workbook.active

def log_result(row_index, success, message=""):
    """
    Log the result of the transaction in column F of the specified row.

    :param row_index: Index of the row to log the result (1-based index)
    :param success: Boolean indicating whether the transaction was successful
    :param message: The message to log (empty for success)
    """
    if success:
        sheet.cell(row=row_index, column=6).value = "Success"
    else:
        sheet.cell(row=row_index, column=6).value = message

def send_transaction(destination_address, amount, asset_code, issuer_address, row_index, min_gas_fee=100):
    try:
        amount = f"{float(amount):.7f}"

        # Reload the source account to get the latest sequence number
        distributor_account = server.load_account(distributor_keypair.public_key)

        # Fetch base fee and ensure it's at least 100
        base_fee = server.fetch_base_fee()
        base_fee = max(base_fee, min_gas_fee)

        # Determine the asset
        if asset_code.upper() == 'XLM':
            asset = Asset.native()
        elif asset_code and issuer_address:
            asset = Asset(asset_code, issuer_address)
        else:
            raise ValueError("Invalid asset info. Must provide 'XLM' or both asset code and issuer address.")

        # Build transaction
        transaction = (
            TransactionBuilder(
                source_account=distributor_account,
                network_passphrase=network_passphrase,
                base_fee=base_fee,
            )
            .append_payment_op(destination=destination_address, amount=str(amount), asset=asset)
            .set_timeout(100)
            .build()
        )

        transaction.sign(distributor_keypair)
        response = server.submit_transaction(transaction)

        if response.get('successful', False):
            print(f"Transaction to {destination_address} for {amount} {asset_code}: Success")
            log_result(row_index, True)
        else:
            error_message = f"Error - {response}"
            print(f"Transaction to {destination_address} for {amount} {asset_code}: {error_message}")
            log_result(row_index, False, error_message)

    except Exception as e:
        if hasattr(e, 'status') and e.status == 504:
            print("504 Gateway Timeout. Retrying...")
            time.sleep(5)  # Delay before retrying
            send_transaction(destination_address, amount, asset_code, issuer_address, row_index)
        elif (
			hasattr(e, 'extras') and 
			e.extras is not None and 
			isinstance(e.extras.get('result_codes'), dict) and 
			e.extras['result_codes'].get('transaction') == 'tx_bad_seq'
		):
            print("Bad sequence number. Reloading account and retrying...")
            time.sleep(1)  # Brief delay before retrying
            send_transaction(destination_address, amount, asset_code, issuer_address, row_index)
        elif (
			hasattr(e, 'extras') and 
			e.extras is not None and 
			isinstance(e.extras.get('result_codes'), dict) and 
			e.extras['result_codes'].get('transaction') == 'tx_too_late'
		):
            print("Transaction time out. Retrying...")
            time.sleep(1)  # Brief delay before retrying
            send_transaction(destination_address, amount, asset_code, issuer_address, row_index)
        elif (
			hasattr(e, 'extras') and 
			e.extras is not None and 
			isinstance(e.extras.get('result_codes'), dict) and 
			e.extras['result_codes'].get('transaction') == 'tx_insufficient_fee'
		):
            if min_gas_fee < 2000:
                print("Insufficient fee. Retrying with "+ str(2 * min_gas_fee) +" Stroops...")
                time.sleep(1)  # Brief delay before retrying
                send_transaction(destination_address, amount, asset_code, issuer_address, row_index, 2 * min_gas_fee )
            else:
                print("Network is too busy at this time. Please try again this transaction at further time.")
                error_message = "Network is too busy at this time. Please try again this transaction at further time."
                log_result(row_index, False, error_message)
        else:
            error_message = f"Transaction failed: {e}"
            print(error_message)
            log_result(row_index, False, error_message)

# Begin processing
print("Now starting transactions")

# Iterate over rows: col A = destination, B = amount, C = asset_code, D = issuer_address
for row_index, row in enumerate(sheet.iter_rows(min_row=1, max_col=4, values_only=True), start=1):
    destination_address, amount, asset_code, issuer_address = row

    # Stop the loop if the destination address is empty
    if not destination_address:
        break

    send_transaction(destination_address, amount, asset_code, issuer_address, row_index)

# Save the updated Excel file
workbook.save('database.xlsx')

print("All transactions processed.")
